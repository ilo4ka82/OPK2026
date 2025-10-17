"""
Обработчик экспорта данных посещаемости в Excel.
"""
import logging
import os
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    ContextTypes,
    ConversationHandler,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    filters,
)
from telegram.constants import ParseMode
import pytz

from ...database import get_attendance_data_for_period
from ...utils import (
    admin_required,
    validate_datetime_format,
    AdminMessages,
    PREDEFINED_SECTORS,
)
from ...keyboards import get_sector_selection_keyboard, get_period_selection_keyboard
from ...states import ExportStates

logger = logging.getLogger(__name__)
MOSCOW_TZ = pytz.timezone('Europe/Moscow')


@admin_required
async def export_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Начинает диалог экспорта данных.
    Команда: /export_attendance
    """
    admin_user_id = update.effective_user.id
    logger.info(f"Админ {admin_user_id} начал процесс экспорта посещаемости.")

    reply_markup = get_sector_selection_keyboard()
    
    await update.message.reply_text(
        AdminMessages.EXPORT_SELECT_SECTOR,
        reply_markup=reply_markup
    )
    
    return ExportStates.SELECT_SECTOR


async def export_sector_selected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Обрабатывает выбор сектора."""
    query = update.callback_query
    await query.answer()

    callback_data_str = query.data

    if callback_data_str == "export_cancel_dialog":
        await query.edit_message_text("Экспорт отменен.")
        context.user_data.clear()
        return ConversationHandler.END

    # Извлекаем ключ сектора
    sector_key = callback_data_str.replace("export_sector_", "")
    context.user_data['export_sector'] = sector_key
    
    logger.info(f"Выбран сектор для экспорта: {sector_key}")

    reply_markup = get_period_selection_keyboard()
    await query.edit_message_text(
        AdminMessages.EXPORT_SELECT_PERIOD,
        reply_markup=reply_markup
    )
    
    return ExportStates.SELECT_PERIOD


async def export_period_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор периода."""
    query = update.callback_query
    await query.answer()

    callback_data_str = query.data

    if callback_data_str == "export_cancel_dialog":
        await query.edit_message_text("Экспорт отменен.")
        context.user_data.clear()
        return ConversationHandler.END

    if callback_data_str == "export_back_to_sector_selection":
        reply_markup = get_sector_selection_keyboard()
        await query.edit_message_text(
            AdminMessages.EXPORT_SELECT_SECTOR,
            reply_markup=reply_markup
        )
        return ExportStates.SELECT_SECTOR

    period_key = callback_data_str.replace("export_period_", "")
    
    # Если произвольный период
    if period_key == "custom":
        await query.edit_message_text(
            "Введите начальную дату периода в формате ДД.ММ.ГГГГ\n"
            "Например: 01.01.2025"
        )
        return ExportStates.GET_START_DATE

    # Иначе вычисляем период
    context.user_data['export_period'] = period_key
    start_date, end_date = _calculate_period_dates(period_key)
    
    if not start_date or not end_date:
        await query.edit_message_text("Ошибка при вычислении дат периода.")
        context.user_data.clear()
        return ConversationHandler.END

    context.user_data['export_start_date'] = start_date
    context.user_data['export_end_date'] = end_date

    # Подтверждение
    sector_key = context.user_data.get('export_sector', 'ALL')
    period_display = _get_period_display_name(period_key)
    
    confirmation_text = AdminMessages.EXPORT_CONFIRM.format(
        sector=_get_sector_display_name(sector_key),
        period=period_display
    )
    
    keyboard = [
        [InlineKeyboardButton("✅ Да, сформировать отчет", callback_data="export_confirm_yes")],
        [InlineKeyboardButton("⬅️ Назад (к выбору периода)", callback_data="export_back_to_period")],
        [InlineKeyboardButton("❌ Отменить экспорт", callback_data="export_cancel_dialog")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(confirmation_text, reply_markup=reply_markup)
    
    return ExportStates.CONFIRM_EXPORT


async def export_custom_start_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получает начальную дату произвольного периода."""
    date_input = update.message.text.strip()
    
    is_valid, dt_obj, error_msg = validate_datetime_format(date_input, '%d.%m.%Y')
    
    if not is_valid:
        await update.message.reply_text(
            f"Неверный формат даты.\n{error_msg}\n\n"
            "Введите начальную дату в формате ДД.ММ.ГГГГ или /cancel для отмены."
        )
        return ExportStates.GET_START_DATE
    
    context.user_data['export_start_date'] = dt_obj
    
    await update.message.reply_text(
        "Теперь введите конечную дату периода в формате ДД.ММ.ГГГГ\n"
        "Например: 31.01.2025"
    )
    
    return ExportStates.GET_END_DATE


async def export_custom_end_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получает конечную дату произвольного периода."""
    date_input = update.message.text.strip()
    
    is_valid, dt_obj, error_msg = validate_datetime_format(date_input, '%d.%m.%Y')
    
    if not is_valid:
        await update.message.reply_text(
            f"Неверный формат даты.\n{error_msg}\n\n"
            "Введите конечную дату в формате ДД.ММ.ГГГГ или /cancel для отмены."
        )
        return ExportStates.GET_END_DATE
    
    start_date = context.user_data.get('export_start_date')
    
    if dt_obj < start_date:
        await update.message.reply_text(
            "Конечная дата не может быть раньше начальной.\n"
            "Введите конечную дату еще раз."
        )
        return ExportStates.GET_END_DATE
    
    context.user_data['export_end_date'] = dt_obj
    context.user_data['export_period'] = 'custom'
    
    # Подтверждение
    sector_key = context.user_data.get('export_sector', 'ALL')
    period_display = f"{start_date.strftime('%d.%m.%Y')} - {dt_obj.strftime('%d.%m.%Y')}"
    
    confirmation_text = AdminMessages.EXPORT_CONFIRM.format(
        sector=_get_sector_display_name(sector_key),
        period=period_display
    )
    
    keyboard = [
        [InlineKeyboardButton("✅ Да, сформировать отчет", callback_data="export_confirm_yes")],
        [InlineKeyboardButton("❌ Отменить экспорт", callback_data="export_cancel_dialog")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(confirmation_text, reply_markup=reply_markup)
    
    return ExportStates.CONFIRM_EXPORT


async def export_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает подтверждение и генерирует отчет."""
    query = update.callback_query
    await query.answer()

    callback_data_str = query.data

    if callback_data_str == "export_cancel_dialog":
        await query.edit_message_text("Экспорт отменен.")
        context.user_data.clear()
        return ConversationHandler.END

    if callback_data_str == "export_back_to_period":
        reply_markup = get_period_selection_keyboard()
        await query.edit_message_text(
            AdminMessages.EXPORT_SELECT_PERIOD,
            reply_markup=reply_markup
        )
        return ExportStates.SELECT_PERIOD

    if callback_data_str == "export_confirm_yes":
        await query.edit_message_text(AdminMessages.EXPORT_GENERATING)
        
        # Получаем данные
        sector_key = context.user_data.get('export_sector', 'ALL')
        start_date = context.user_data.get('export_start_date')
        end_date = context.user_data.get('export_end_date')
        
        if not start_date or not end_date:
            await query.message.reply_text("Ошибка: даты не найдены.")
            context.user_data.clear()
            return ConversationHandler.END
        
        # Получаем данные из БД
        attendance_data = get_attendance_data_for_period(start_date, end_date, sector_key)
        
        if not attendance_data:
            period_display = _get_period_display_name(context.user_data.get('export_period', 'custom'))
            await query.message.reply_text(
                AdminMessages.EXPORT_NO_DATA.format(
                    sector=_get_sector_display_name(sector_key),
                    period=period_display
                )
            )
            context.user_data.clear()
            return ConversationHandler.END
        
        # Генерируем Excel файл
        try:
            file_path = _generate_excel_report(attendance_data, sector_key, start_date, end_date)
            
            # Отправляем файл
            with open(file_path, 'rb') as file:
                await query.message.reply_document(
                    document=file,
                    filename=os.path.basename(file_path),
                    caption=f"📊 Отчет по посещаемости готов!\n"
                            f"Сектор: {_get_sector_display_name(sector_key)}\n"
                            f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
                )
            
            # Удаляем временный файл
            os.remove(file_path)
            logger.info(f"Отчет успешно сгенерирован и отправлен: {file_path}")
            
        except Exception as e:
            logger.error(f"Ошибка при генерации отчета: {e}", exc_info=True)
            await query.message.reply_text("Произошла ошибка при формировании отчета.")
        
        context.user_data.clear()
        return ConversationHandler.END

    return ExportStates.CONFIRM_EXPORT


async def cancel_export_dialog(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Отменяет диалог экспорта."""
    await update.message.reply_text("Экспорт отменен.")
    context.user_data.clear()
    return ConversationHandler.END


# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ===

def _calculate_period_dates(period_key: str):
    """Вычисляет даты начала и конца периода."""
    now = datetime.now(MOSCOW_TZ)
    
    if period_key == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    elif period_key == "yesterday":
        yesterday = now - timedelta(days=1)
        start = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
        end = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    elif period_key == "this_week":
        start = now - timedelta(days=now.weekday())
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    elif period_key == "last_week":
        start = now - timedelta(days=now.weekday() + 7)
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=6, hours=23, minutes=59, seconds=59)
    
    elif period_key == "this_month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    elif period_key == "last_month":
        first_day_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        start = last_day_last_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = last_day_last_month.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    else:
        return None, None
    
    return start, end


def _get_period_display_name(period_key: str) -> str:
    """Возвращает отображаемое название периода."""
    period_names = {
        "today": "Сегодня",
        "yesterday": "Вчера",
        "this_week": "Эта неделя",
        "last_week": "Прошлая неделя",
        "this_month": "Этот месяц",
        "last_month": "Прошлый месяц",
        "custom": "Произвольный период"
    }
    return period_names.get(period_key, period_key)


def _get_sector_display_name(sector_key: str) -> str:
    """Возвращает отображаемое название сектора."""
    if sector_key == "ALL":
        return "Все секторы"
    
    for sector_name in PREDEFINED_SECTORS:
        if sector_name.endswith(sector_key):
            return sector_name
    
    return sector_key


def _generate_excel_report(data: list, sector_key: str, start_date: datetime, end_date: datetime) -> str:
    """
    Генерирует Excel файл с отчетом.
    
    Returns:
        Путь к сгенерированному файлу
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        logger.error("openpyxl не установлен. Установите: pip install openpyxl")
        raise

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"
    
    # Заголовок
    ws['A1'] = "Отчет по посещаемости"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Сектор: {_get_sector_display_name(sector_key)}"
    ws['A3'] = f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    
    # Заголовки таблицы
    headers = ["№", "ФИО", "Username", "Сектор", "Время прихода", "Время ухода", "Длительность (ч)"]
    header_row = 5
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for idx, row_data in enumerate(data, 1):
        row_num = header_row + idx
        
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=row_data.get('application_full_name', 'N/A'))
        ws.cell(row=row_num, column=3, value=row_data.get('username', 'N/A'))
        ws.cell(row=row_num, column=4, value=row_data.get('application_department', 'N/A'))
        
        # Время прихода
        checkin = row_data.get('session_start_time', '')
        ws.cell(row=row_num, column=5, value=checkin)
        
        # Время ухода
        checkout = row_data.get('session_end_time', 'Активна')
        ws.cell(row=row_num, column=6, value=checkout)
        
        # Длительность
        if checkout and checkout != 'Активна':
            try:
                checkin_dt = datetime.strptime(checkin, '%Y-%m-%d %H:%M:%S')
                checkout_dt = datetime.strptime(checkout, '%Y-%m-%d %H:%M:%S')
                duration = (checkout_dt - checkin_dt).total_seconds() / 3600
                ws.cell(row=row_num, column=7, value=f"{duration:.2f}")
            except:
                ws.cell(row=row_num, column=7, value="N/A")
        else:
            ws.cell(row=row_num, column=7, value="Активна")
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    filename = f"attendance_{sector_key}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    file_path = os.path.join("/tmp", filename)
    wb.save(file_path)
    
    return file_path


# ConversationHandler для экспорта
export_conv_handler = ConversationHandler(
    entry_points=[
        CommandHandler("export_attendance", export_start)
    ],
    states={
        ExportStates.SELECT_SECTOR: [
            CallbackQueryHandler(export_sector_selected)
        ],
        ExportStates.SELECT_PERIOD: [
            CallbackQueryHandler(export_period_selected)
        ],
        ExportStates.GET_START_DATE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, export_custom_start_date)
        ],
        ExportStates.GET_END_DATE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, export_custom_end_date)
        ],
        ExportStates.CONFIRM_EXPORT: [
            CallbackQueryHandler(export_confirmation)
        ],
    },
    fallbacks=[
        CommandHandler("cancel", cancel_export_dialog)
    ],
    name="export_attendance_flow",
    per_user=True,
    per_chat=True,
    allow_reentry=True
)